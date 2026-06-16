import io
import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def df_to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Data",
                      title: str = "", report_curr: str = "USD") -> bytes:
    """
    Convierte un DataFrame en un Excel formateado estilo Family Office:
    - Header azul oscuro, texto blanco, negrita
    - Filas alternadas gris muy claro / blanco
    - Columnas de dinero con formato #,##0
    - Columnas % con formato 0.00%
    - Columnas x (múltiplos) con formato 0.00x
    - Anchos automáticos
    - Fila de título si se pasa `title`
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    # ── Paleta ────────────────────────────────────────────────────────────────
    HDR_FILL  = PatternFill("solid", fgColor="002060")   # azul oscuro
    ALT_FILL  = PatternFill("solid", fgColor="EEF2FA")   # gris muy claro
    TTL_FILL  = PatternFill("solid", fgColor="0D1929")   # azul marino (título)
    WHT_FONT  = Font(name="Arial", color="FFFFFF", bold=True, size=10)
    HDR_FONT  = Font(name="Arial", color="FFFFFF", bold=True, size=10)
    BODY_FONT = Font(name="Arial", color="1A1A2E", size=10)
    TTL_FONT  = Font(name="Arial", color="FFFFFF", bold=True, size=12)
    THIN      = Side(style="thin", color="D0D8E8")
    BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    HDR_BRD   = Border(left=THIN, right=THIN,
                        top=Side(style="medium", color="002060"),
                        bottom=Side(style="medium", color="002060"))

    curr_sym  = {"USD": "$", "EUR": "€", "GBP": "£"}.get(report_curr, "$")
    n_cols    = len(df.columns) + 1   # +1 por el índice

    row_offset = 1

    # ── Fila de título ────────────────────────────────────────────────────────
    if title:
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=n_cols)
        cell = ws.cell(row=1, column=1, value=title)
        cell.font      = TTL_FONT
        cell.fill      = TTL_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center",
                                   indent=1)
        ws.row_dimensions[1].height = 22
        row_offset = 2

    # ── Header ────────────────────────────────────────────────────────────────
    hdr_row = row_offset
    ws.cell(row=hdr_row, column=1, value="#").font   = HDR_FONT
    ws.cell(row=hdr_row, column=1).fill              = HDR_FILL
    ws.cell(row=hdr_row, column=1).alignment         = Alignment(horizontal="center")
    ws.cell(row=hdr_row, column=1).border            = HDR_BRD
    ws.row_dimensions[hdr_row].height = 18

    for ci, col in enumerate(df.columns, start=2):
        c = ws.cell(row=hdr_row, column=ci, value=str(col))
        c.font      = HDR_FONT
        c.fill      = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border    = HDR_BRD

    # ── Detectar tipo de columna para formato ─────────────────────────────────
    def col_fmt(col_name: str):
        cn = str(col_name).lower()
        # Cualquier columna con % en el nombre → formato porcentaje literal
        if cn.startswith('%') or ' %' in cn or cn.endswith('%') or 'irr' in cn or 'rent' in cn:
            return '0.00"%"', "right"
        if any(x in cn for x in ["tvpi", "dpi", "rvpi", "múltiplo", "multiple"]):
            return '0.00"x"', "right"
        if any(x in cn for x in ["commit", "paid", "unfunded", "distributed",
                                   "nav", "total value", "utilidad", "ganancia",
                                   "impuesto", "calls", "distribuciones",
                                   "net cash", "value", "(mm)"]):
            return "#,##0.0", "right"
        if cn in ["vintage", "n° fondos", "fondos", "año", "# inv", "duration"]:
            return "0.0", "center"
        return "General", "left"

    # ── Filas de datos ────────────────────────────────────────────────────────
    for ri, (idx, row) in enumerate(df.iterrows(), start=1):
        xr = hdr_row + ri
        fill = ALT_FILL if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")

        # Índice
        c = ws.cell(row=xr, column=1, value=idx)
        c.font = BODY_FONT; c.fill = fill; c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")

        for ci, col in enumerate(df.columns, start=2):
            val = row[col]
            fmt, align = col_fmt(col)

            # Limpiar valores formateados como string (ej. "1.33x", "12.58%")
            if isinstance(val, str):
                clean = val.replace("x","").replace("%","").replace(",","").strip()
                try:
                    val = float(clean)
                    # NO dividir por 100 — los IRR ya vienen en escala correcta
                    # (ej. 12.58 significa 12.58%, el formato '0.00"%"' lo muestra bien)
                except ValueError:
                    pass  # dejar como string

            c = ws.cell(row=xr, column=ci, value=val)
            c.font      = BODY_FONT
            c.fill      = fill
            c.border    = BORDER
            c.alignment = Alignment(horizontal=align, vertical="center")
            if fmt != "General":
                c.number_format = fmt

        ws.row_dimensions[xr].height = 16

    # ── Anchos automáticos ────────────────────────────────────────────────────
    for ci in range(1, n_cols + 1):
        max_len = 0
        col_letter = get_column_letter(ci)
        for row_cells in ws.iter_rows(min_col=ci, max_col=ci):
            for cell in row_cells:
                try:
                    cell_len = len(str(cell.value)) if cell.value else 0
                    max_len  = max(max_len, cell_len)
                except:
                    pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 30)

    # ── Freeze panes (fijar header) ───────────────────────────────────────────
    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=2)

    # ── Guardar en bytes ──────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def excel_download_btn(df: pd.DataFrame, label: str, filename: str,
                       sheet_name: str, title: str, report_curr: str,
                       key: str):
    """Botón de descarga compacto con ícono, alineado a la derecha."""
    xlsx_bytes = df_to_excel_bytes(df, sheet_name=sheet_name,
                                    title=title, report_curr=report_curr)
    # Columna vacía + botón para alinearlo a la derecha
    _, btn_col = st.columns([6, 1])
    with btn_col:
        st.download_button(
            label="⬇ .xlsx",
            data=xlsx_bytes,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=key,
            help=f"Descargar {label} en Excel",
            use_container_width=True,
        )


# ─────────────────────────────────────────────────────────────────────────────